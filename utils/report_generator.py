import os
import pandas as pd
import numpy as np
from datetime import datetime

# Import ReportLab modules
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Import OpenPyXL modules for Excel formatting
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_pdf_report(metrics_dict, coef_df, risk_counts, revenue_impact, output_path):
    """
    Generates a presentation-ready executive PDF report summarizing the churn predictions,
    model performance, risk segmentation, and expected revenue impact.
    """
    # Create the document
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        rightMargin=54,
        leftMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    primary_color = colors.HexColor("#1E3A8A")   # Dark Navy
    secondary_color = colors.HexColor("#3B82F6") # Blue
    accent_color = colors.HexColor("#EF4444")    # Crimson Red
    text_color = colors.HexColor("#1F2937")      # Charcoal
    bg_light = colors.HexColor("#F3F4F6")        # Light Grey
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=primary_color,
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=30
    )
    
    h1_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=primary_color,
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    h2_style = ParagraphStyle(
        'SubSectionHeading',
        parent=styles['Heading3'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=secondary_color,
        spaceBefore=10,
        spaceAfter=6,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=text_color,
        spaceAfter=8
    )
    
    bullet_style = ParagraphStyle(
        'BulletCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=text_color,
        leftIndent=20,
        firstLineIndent=-10,
        spaceAfter=4
    )
    
    kpi_title_style = ParagraphStyle(
        'KPITitle',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#4B5563"),
        alignment=1 # Center
    )
    
    kpi_value_style = ParagraphStyle(
        'KPIValue',
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=primary_color,
        alignment=1 # Center
    )
    
    story = []
    
    # --- HEADER SECTION ---
    story.append(Paragraph("Predictive Churn Analysis Platform", title_style))
    date_str = datetime.now().strftime("%B %d, %Y")
    story.append(Paragraph(f"<b>Executive Summary Report</b> &nbsp;|&nbsp; Generated on: {date_str} &nbsp;|&nbsp; Class: Subscription Analytics", subtitle_style))
    story.append(Spacer(1, 10))
    
    # --- SECTION 1: BUSINESS KPIs ---
    story.append(Paragraph("Business Retention KPIs", h1_style))
    story.append(Paragraph("A summary of predicted customer retention and risk levels based on the trained predictive model:", body_style))
    story.append(Spacer(1, 8))
    
    # Define KPI Grid Table
    total_cust = revenue_impact['total_customers']
    renewal_rate = revenue_impact['expected_renewal_rate'] * 100
    rev_retained = revenue_impact['expected_revenue_retained']
    rev_at_risk = revenue_impact['expected_revenue_at_risk']
    
    kpi_data = [
        [
            Paragraph("TOTAL CUSTOMERS", kpi_title_style),
            Paragraph("EXPECTED RENEWAL RATE", kpi_title_style),
            Paragraph("EXPECTED MONTHLY REVENUE RETAINED", kpi_title_style),
            Paragraph("MONTHLY REVENUE AT RISK", kpi_title_style)
        ],
        [
            Paragraph(f"{total_cust:,}", kpi_value_style),
            Paragraph(f"{renewal_rate:.1f}%", kpi_value_style),
            Paragraph(f"${rev_retained:,.2f}", kpi_value_style),
            Paragraph(f"${rev_at_risk:,.2f}", ParagraphStyle('KPIValueRisk', parent=kpi_value_style, textColor=accent_color))
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[1.25*inch, 1.75*inch, 2.25*inch, 1.75*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), bg_light),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TEXTCOLOR', (0,0), (-1,-1), text_color),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.white),
        ('BOX', (0,0), (-1,-1), 1, primary_color),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # --- SECTION 2: RISK SEGMENTATION ---
    story.append(Paragraph("Customer Risk Segmentation", h1_style))
    story.append(Paragraph("Customers have been segmented into three risk categories based on their predicted renewal probability:", body_style))
    
    # Risk table
    high_risk = risk_counts.get('High Churn Risk 🔴', 0)
    med_risk = risk_counts.get('Medium Risk 🟡', 0)
    low_risk = risk_counts.get('Likely Renewal 🟢', 0)
    
    risk_data = [
        ["Risk Tier", "Renewal Probability", "Customer Count", "Percentage of Customer Base"],
        ["🔴 High Churn Risk", "0% - 30%", f"{high_risk:,}", f"{(high_risk/total_cust)*100:.1f}%"],
        ["🟡 Medium Risk", "31% - 70%", f"{med_risk:,}", f"{(med_risk/total_cust)*100:.1f}%"],
        ["🟢 Likely Renewal", "71% - 100%", f"{low_risk:,}", f"{(low_risk/total_cust)*100:.1f}%"]
    ]
    
    risk_table = Table(risk_data, colWidths=[2.0*inch, 1.5*inch, 1.5*inch, 2.0*inch])
    risk_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(risk_table)
    story.append(Spacer(1, 15))
    
    # --- SECTION 3: MODEL PERFORMANCE ---
    story.append(Paragraph("Model Evaluation Metrics", h1_style))
    story.append(Paragraph("The primary predictive model is trained using <b>Logistic Regression</b>. The following classification performance metrics were achieved on the holdout test set:", body_style))
    
    metrics_data = [
        ["Evaluation Metric", "Score", "Performance Interpretation"],
        ["Model Accuracy", f"{metrics_dict['accuracy']*100:.1f}%", "Overall percentage of correct predictions."],
        ["Precision Score", f"{metrics_dict['precision']*100:.1f}%", "Percentage of users predicted to renew who actually did."],
        ["Recall (Sensitivity)", f"{metrics_dict['recall']*100:.1f}%", "Percentage of actual renewers correctly identified by model."],
        ["F1 Score", f"{metrics_dict['f1_score']*100:.1f}%", "Harmonic mean of precision and recall."],
        ["ROC-AUC", f"{metrics_dict['roc_auc']:.3f}", "Ability of model to distinguish between classes (1 = perfect)."]
    ]
    
    metrics_table = Table(metrics_data, colWidths=[2.0*inch, 1.25*inch, 3.75*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), secondary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 20))
    
    # Force page break for remaining sections to look clean
    story.append(PageBreak())
    
    # --- SECTION 4: FEATURE IMPORTANCE ---
    story.append(Paragraph("Top Factors Driving Renewal Decisions", h1_style))
    story.append(Paragraph("The table below shows the features impacting customer renewals. <i>Positive coefficients</i> drive renewals (customers stay), while <i>negative coefficients</i> increase churn risk (customers leave).", body_style))
    
    coef_table_data = [["Rank", "Feature Name", "Impact Coefficient", "Odds Ratio", "P-Value", "Statistical Significance"]]
    
    # Map p-value to text
    for i, (_, row) in enumerate(coef_df.iterrows()):
        p_val = row['P-Value']
        sig_str = "Highly Significant" if p_val < 0.01 else ("Significant" if p_val < 0.05 else "Not Significant")
        coef_table_data.append([
            str(i+1),
            row['Feature'].replace('_', ' '),
            f"{row['Coefficient']:.4f}",
            f"{row['Odds_Ratio']:.3f}x",
            f"{p_val:.4f}",
            sig_str
        ])
        
    coef_table = Table(coef_table_data, colWidths=[0.6*inch, 2.2*inch, 1.2*inch, 0.9*inch, 0.9*inch, 1.2*inch])
    coef_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, bg_light]),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (2,0), (-2,-1), 'RIGHT'),
        ('ALIGN', (-1,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(coef_table)
    story.append(Spacer(1, 15))
    
    # --- SECTION 5: RECOMMENDATIONS ---
    story.append(Paragraph("Actionable Customer Retention Recommendations", h1_style))
    story.append(Paragraph("Based on the churn drivers and risk levels, we recommend deploying the following programmatic retention campaigns:", body_style))
    
    story.append(Paragraph("<b>1. High Churn Risk Cohort (0% - 30% Probability):</b>", h2_style))
    story.append(Paragraph("• <b>Immediate Outreach</b>: Trigger direct customer success calls within 24 hours of falling into the High Risk category.", bullet_style))
    story.append(Paragraph("• <b>Financial Incentive</b>: Offer a localized loyalty discount (e.g., 20% off for 6 months) or a free renewal month in exchange for extending contract terms.", bullet_style))
    story.append(Paragraph("• <b>Product Re-engagement</b>: Send personalized newsletters highlighting features the customer hasn't used yet based on their low product usage.", bullet_style))
    
    story.append(Paragraph("<b>2. Medium Risk Cohort (31% - 70% Probability):</b>", h2_style))
    story.append(Paragraph("• <b>Targeted Campaigns</b>: Send structured email/SMS engagement campaigns that showcase value, premium trials, or testimonials.", bullet_style))
    story.append(Paragraph("• <b>Support Ticket Resolution</b>: Prioritize unresolved customer complaints for users with high support ticket counts.", bullet_style))
    story.append(Paragraph("• <b>NPS Probing</b>: Send a Net Promoter Score survey to determine underlying issues before their renewal date.", bullet_style))
    
    story.append(Paragraph("<b>3. Likely Renewal Cohort (71% - 100% Probability):</b>", h2_style))
    story.append(Paragraph("• <b>Upsell Opportunities</b>: Target high-engagement users with premium plan features or add-ons.", bullet_style))
    story.append(Paragraph("• <b>Referral Program</b>: Encourage these loyal users to advocate for your brand by providing referral credits or discounts.", bullet_style))
    story.append(Spacer(1, 10))
    
    # --- DISCLAIMER / SIGN OFF ---
    story.append(Spacer(1, 20))
    disclaimer_style = ParagraphStyle(
        'DisclaimerText',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#7F8C8D"),
        alignment=1 # Center
    )
    story.append(Paragraph("Disclaimer: This predictive report contains simulations and estimates based on statistical modeling. Decisions should be combined with qualitative business judgement.", disclaimer_style))
    
    # Build Document
    doc.build(story)

def generate_excel_report(scored_df, output_path):
    """
    Exports a styled Excel report of scored customers, including risk tiers
    and engineered metrics, formatting the spreadsheet for immediate business use.
    """
    # Exclude technical internal arrays/columns if any exist
    cols_to_export = [
        'Customer_ID', 'Age', 'Tenure', 'Monthly_Spend', 'Login_Frequency',
        'Session_Duration', 'Support_Tickets', 'Product_Usage', 'Last_Login_Days',
        'Subscription_Type', 'Renewal_Status', 'Engagement_Score', 'Support_Risk',
        'Recency_Score', 'Customer_Health_Score', 'Renewal_Probability', 'Risk_Level'
    ]
    
    # Subset only columns that actually exist
    existing_cols = [c for c in cols_to_export if c in scored_df.columns]
    df_export = scored_df[existing_cols].copy()
    
    # Rename for professional look
    df_export = df_export.rename(columns={
        'Customer_ID': 'Customer ID',
        'Monthly_Spend': 'Monthly Spend ($)',
        'Login_Frequency': 'Monthly Logins',
        'Session_Duration': 'Avg Session Duration (Mins)',
        'Support_Tickets': 'Support Tickets',
        'Product_Usage': 'Product Usage Score',
        'Last_Login_Days': 'Days Since Last Login',
        'Subscription_Type': 'Subscription Plan',
        'Renewal_Status': 'Actual Renewal Status',
        'Engagement_Score': 'Engagement Score',
        'Support_Risk': 'Support Ticket Rate',
        'Recency_Score': 'Recency (Days)',
        'Customer_Health_Score': 'Customer Health Score',
        'Renewal_Probability': 'Renewal Probability (%)',
        'Risk_Level': 'Risk Level'
    })
    
    # Save using Excel Writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Customer Predictions', index=False)
        
        # Access openpyxl worksheet and workbook
        workbook = writer.book
        worksheet = writer.sheets['Customer Predictions']
        
        # Styles
        navy_header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=10)
        bold_font = Font(name='Arial', size=10, bold=True)
        
        # Border
        thin_border_side = Side(style='thin', color='D1D5DB')
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Conditional formatting fills for risk
        red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid') # Red
        yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid') # Yellow
        green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid') # Green
        
        # Apply header styling
        for col_idx, col_name in enumerate(df_export.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = navy_header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet.row_dimensions[1].height = 28
            
        # Apply data formatting
        for row_idx in range(2, len(df_export) + 2):
            worksheet.row_dimensions[row_idx].height = 18
            for col_idx, col_name in enumerate(df_export.columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignments
                if col_name in ['Customer ID', 'Subscription Plan', 'Risk Level']:
                    cell.alignment = Alignment(horizontal='center')
                elif col_name in ['Actual Renewal Status']:
                    cell.alignment = Alignment(horizontal='center')
                    # format actual status as binary indicator
                    cell.number_format = '0'
                elif col_name in ['Monthly Spend ($)', 'Renewal Probability (%)']:
                    cell.alignment = Alignment(horizontal='right')
                    if col_name == 'Monthly Spend ($)':
                        cell.number_format = '$#,##0.00'
                    else:
                        # Convert fraction to percentage representation in Excel
                        # (assumes raw prediction columns are floats between 0 and 1)
                        cell.value = float(cell.value)
                        cell.number_format = '0.0%'
                else:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0' if isinstance(cell.value, (int, np.integer)) else '#,##0.0'
                    
                # Apply Risk color highlights
                if col_name == 'Risk Level':
                    val = str(cell.value)
                    if 'High Churn Risk' in val or '🔴' in val:
                        cell.fill = red_fill
                        cell.font = bold_font
                    elif 'Medium Risk' in val or '🟡' in val:
                        cell.fill = yellow_fill
                        cell.font = bold_font
                    elif 'Likely Renewal' in val or '🟢' in val:
                        cell.fill = green_fill
                        cell.font = bold_font
                        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row == 1:
                    # Header values can be long, so split them to approximate widths
                    max_len = max(max_len, len(val_str.split('\n')[0]))
                else:
                    max_len = max(max_len, len(val_str))
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Freeze top row
        worksheet.freeze_panes = 'A2'
